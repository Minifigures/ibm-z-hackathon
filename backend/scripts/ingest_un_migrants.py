"""Extract 2020 bilateral migrant stocks from UN DESA International Migrant
Stock dataset for the 71 countries in countries.json.

Source: UN DESA Population Division, "International Migrant Stock 2020:
Destination and origin" (Table 1, both-sexes 2020 column). Migrant stock
captures the cumulative diaspora population (e.g., 12M Mexicans living in
USA in 2020) which is a strong predictor of bilateral travel volume that
gravity models miss.

Output: backend/app/data/un_migrant_stock.json with the schema::

    {
        "_meta": { ... },
        "year": 2020,
        "stocks": { "MEX_USA": 11900000, ... }   # origin_destination
    }

Run from backend/::
    python -m scripts.ingest_un_migrants
"""

from __future__ import annotations

import json
import logging
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

REPO_BACKEND = Path(__file__).resolve().parent.parent
SRC = REPO_BACKEND / "scripts" / "data" / "un_migrant_stock_2020.xlsx"
M49_MAP = REPO_BACKEND / "scripts" / "data" / "m49_to_iso3.json"
OUT = REPO_BACKEND / "app" / "data" / "un_migrant_stock.json"

YEAR_2020_COL = 13   # row-10 header: col13 = 2020 (both-sexes block)
DEST_NAME_COL = 1
DEST_CODE_COL = 3
ORIG_NAME_COL = 5
ORIG_CODE_COL = 6
DATA_START_ROW = 12  # row 11 is the WORLD/WORLD aggregate; data follows


def main():
    if not SRC.exists():
        raise SystemExit(f"Source not found: {SRC}")
    if not M49_MAP.exists():
        raise SystemExit(f"M49 map not found: {M49_MAP}")

    m49_to_iso3 = {int(k): v for k, v in json.loads(M49_MAP.read_text()).items()}
    valid_codes = set(m49_to_iso3.keys())
    logger.info("Loading workbook (this may take ~30s for 37k rows)...")
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["Table 1"]

    stocks: dict[tuple[str, str], int] = {}
    rows_seen = 0
    rows_kept = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        rows_seen += 1
        try:
            dest_code = row[DEST_CODE_COL]
            orig_code = row[ORIG_CODE_COL]
        except IndexError:
            continue
        if dest_code is None or orig_code is None:
            continue
        try:
            dc = int(dest_code)
            oc = int(orig_code)
        except (TypeError, ValueError):
            continue
        if dc not in valid_codes or oc not in valid_codes:
            continue
        if dc == oc:
            continue
        val = row[YEAR_2020_COL]
        if val is None or val == "..":
            continue
        try:
            v = int(val)
        except (TypeError, ValueError):
            continue
        if v <= 0:
            continue
        dest_iso = m49_to_iso3[dc]
        orig_iso = m49_to_iso3[oc]
        stocks[(orig_iso, dest_iso)] = v
        rows_kept += 1

    logger.info("rows seen=%d, country-pair rows kept=%d, unique pairs=%d",
                rows_seen, rows_kept, len(stocks))

    sorted_stocks = sorted(stocks.items(), key=lambda x: -x[1])
    payload = {
        "_meta": {
            "source": "UN DESA Population Division, International Migrant Stock 2020 (Table 1, both-sexes 2020 column)",
            "url": "https://www.un.org/development/desa/pd/content/international-migrant-stock",
            "year": 2020,
            "n_pairs": len(stocks),
            "note": "Bilateral migrant stocks (diaspora populations). Key 'origin_destination' (e.g., MEX_USA = number of Mexican-born people living in USA in 2020). Used as a proxy for diaspora-driven travel flows that pure gravity models systematically under-predict.",
        },
        "year": 2020,
        "stocks": {f"{a}_{b}": v for (a, b), v in sorted_stocks},
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2))
    logger.info("Wrote %s", OUT)

    print("\nTop 15 bilateral migrant stocks (origin -> destination, 2020):")
    for (a, b), v in sorted_stocks[:15]:
        print(f"  {a} -> {b}: {v:>11,d}")


if __name__ == "__main__":
    main()
